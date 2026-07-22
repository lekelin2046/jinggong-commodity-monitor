#!/usr/bin/env python3
"""
全品种每日抓取脚本（16项）
覆盖: ccmn(7) + SMM(7) + 钨粉(1) + WTI(1) = 16

用法: python3 daily_update_all.py
"""

import sys, os, json, datetime, asyncio, re, subprocess, time
from pathlib import Path
from typing import Optional

SCRIPT_DIR = Path(__file__).parent
sys.path.insert(0, str(SCRIPT_DIR))

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed", file=sys.stderr)
    sys.exit(1)

# ===== 常量 =====
# 2026-07-08：修正数据源为原始 Excel（共享(2).xlsx 为不完整副本）
EXCEL_PATH = SCRIPT_DIR / "2026年有色金属市场价格.xlsx"
SHEET_NAME = "日均价（2026年市场）"
TODAY = datetime.date.today().isoformat()

# Excel 列 → (品种代码, 来源)
COL_MAP = {
    2:  ("ADC12",    "smm"),  3:  ("A380",      "smm"),
    4:  ("AlSi9Cu3", "smm"),  5:  ("A356",      "smm"),
    6:  ("A00_AL",   "ccmn"), 7:  ("CU",        "ccmn"),
    8:  ("SI_441",   "ccmn"), 9:  ("SI_3303",   "ccmn"),
    10: ("MG",       "ccmn"), 11: ("MN",        "ccmn"),
    12: ("SI_331",   "ccmn"),
    13: ("Wenxi_MG", "asianmetal"),  14: ("AM60B",     "smm"),
    15: ("AZ91D",    "smm"),
    16: ("W",        "web"),  17: ("WTI",       "wti"),
    18: ("IRON_ORE", "steel"), 19: ("COKE",     "steel"),
    # 2026-07-21 新增：卓创 6 品种 + SMM ADC12日本CIF
    20: ("SS_304",   "sci99"), 21: ("SS_409",     "sci99"),
    22: ("SS_439",   "sci99"), 23: ("SS_441",     "sci99"),
    24: ("NICKEL_IRON", "sci99"), 25: ("HIGH_CARBON_FECR", "sci99"),
    26: ("ADC12_JAPAN_CIF", "smm"),
}

# CCMN 返回 key → 品种代码
CCMN_KEYS = {
    "A00_AL":       "A00_AL",
    "CU":           "CU",
    "SI_553_331_AVG":  "SI_441",
    "SI_3303_2202_MIN": "SI_3303",
    "SI_553_331_MAX":  "SI_331",
    "MG":           "MG",
    "MN":           "MN",
}


def get_row(ws):
    """查找或创建今天的数据行"""
    td = datetime.date.today()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, datetime.datetime) and v.date() == td:
            return r
    nr = ws.max_row + 1
    ws.cell(nr, 1).value = td
    return nr


# ===== CCMN 抓取 =====
async def fetch_ccmn() -> dict:
    print("  ccmn 长江现货...", end="", flush=True)
    from jinggong_monitor.fetcher_ccmn import CcmnFetcher
    fetcher = CcmnFetcher()
    raw = fetcher.fetch()
    result = {}
    for ccmn_key, code in CCMN_KEYS.items():
        if ccmn_key in raw and raw[ccmn_key] is not None:
            result[code] = raw[ccmn_key]
    print(f" {len(result)} 品种")
    return result


# ===== SMM 抓取 =====
async def fetch_smm() -> dict:
    print("  SMM 上海有色...", end="", flush=True)
    from jinggong_monitor.fetcher_smm import _fetch_smm_raw
    try:
        prices = await asyncio.wait_for(_fetch_smm_raw(), timeout=90)
    except asyncio.TimeoutError:
        print(" 超时（90s）")
        return {}
    mapped = {}
    for k, v in prices.items():
        mapped["Wenxi_MG" if k == "WenxiMG" else k] = v
    # ADC12_JAPAN_CIF 直接透传
    if "ADC12_JAPAN_CIF" in prices:
        mapped["ADC12_JAPAN_CIF"] = prices["ADC12_JAPAN_CIF"]
    print(f" {len(mapped)} 品种")
    return mapped


# ===== 亚洲金属网闻喜镁锭 =====
async def fetch_asianmetal() -> dict:
    print("  亚洲金属网（闻喜镁锭）...", end="", flush=True)
    try:
        from jinggong_monitor.fetcher_asianmetal import fetch_async
        prices = await asyncio.wait_for(fetch_async(), timeout=120)
        if prices and "Wenxi_MG" in prices:
            print(f" {prices['Wenxi_MG']} 元/吨")
            return prices
        print(" 未获取")
        return {}
    except asyncio.TimeoutError:
        print(f" 超时（120s）")
        return {}
    except Exception as e:
        print(f" 失败: {e}")
        return {}


# ===== 钨粉 =====
def fetch_tungsten() -> dict:
    """中钨在线钨粉价格（HTTP 抓取，规避 www.chinatungsten.com 的 HTTPS/SSL 故障）

    说明：此前直接 requests.get("https://www.chinatungsten.com/price/")，但该站点
    HTTPS 证书/TLS 已损坏（SSLError: record layer failure），稳定失败。
    正确入口是 http://news.chinatungsten.com（英文门户的中文每日价栏目），由
    ChinatungstenFetcher 解析「钨粉价格 X 元/千克」。2026-07-08 实测可正常返回。
    """
    print("  中钨在线（钨粉）...", end="", flush=True)
    try:
        from jinggong_monitor.fetcher_tungsten import ChinatungstenFetcher
        res = ChinatungstenFetcher().fetch()
        if res and "W" in res:
            v = float(res["W"])
            print(f" {v} 元/千克")
            return {"W": v}
        print(" 未匹配到")
        return {}
    except Exception as e:
        print(f" 失败: {e}")
        return {}


# ===== WTI =====
def fetch_wti() -> dict:
    """获取WTI估算（CL实时价 or SC原油/7.15）"""
    print("  WTI 原油...", end="", flush=True)
    try:
        import akshare as ak
        df = ak.futures_foreign_commodity_realtime(symbol="CL")
        price = round(float(df.iloc[0]["最新价"]), 2)
        print(f" {price} USD (CL)")
        return {"WTI": price}
    except Exception:
        pass
    try:
        import akshare as ak
        dt = TODAY.replace("-", "")
        df = ak.futures_main_sina(symbol="SC0", start_date=dt, end_date=dt)
        close = float(df.iloc[-1]["收盘价"])
        price = round(close / 7.15, 2)
        print(f" {price} USD (SC/7.15)")
        return {"WTI": price}
    except Exception as e:
        print(f" 失败: {e}")
        return {}


# ===== 钢铁（铁矿石 / 冶金焦）=====
async def fetch_steel() -> dict:
    """SMM 钢铁频道：进口铁矿石卡粉65%京唐港、一级冶金焦 MT<7 全国均价

    无需登录；Playwright 渲染 + 正则提取（见 jinggong_monitor.fetcher_steel）。
    在本脚本已运行的事件循环中执行，必须走 fetch_async() 避免 asyncio.run 嵌套。
    """
    print("  钢铁（铁矿石/冶金焦）...", end="", flush=True)
    try:
        from jinggong_monitor.fetcher_steel import fetch_async
        res = await fetch_async()
        if res:
            print(f" 铁矿石={res.get('IRON_ORE')} 冶金焦={res.get('COKE')}")
            return res
        print(" 未匹配到")
        return {}
    except Exception as e:
        print(f" 失败: {e}")
        return {}


# ===== 卓创资讯（6 品种）=====
async def fetch_sci99() -> dict:
    """卓创资讯：Playwright 加载 cookie 抓取 6 个品种平均价
    依赖 cookies/sci99.json，cookie 失效时跳过（不估算，守铁律）
    """
    print("  卓创资讯（6品种）...", end="", flush=True)
    try:
        from jinggong_monitor.fetcher_sci99 import fetch_sci99_async
        prices = await asyncio.wait_for(fetch_sci99_async(), timeout=120)
        if prices:
            names = {"SS_304": "304", "SS_409": "409", "SS_439": "439",
                     "SS_441": "441", "NICKEL_IRON": "镍铁", "HIGH_CARBON_FECR": "铬铁"}
            summary = " ".join(f"{names.get(k,k)}={v}" for k,v in sorted(prices.items()))
            print(f" {summary}")
            return prices
        print(" 无结果（cookie 可能过期）")
        return {}
    except asyncio.TimeoutError:
        print(f" 超时（120s）")
        return {}
    except Exception as e:
        print(f" 失败: {e}")
        import traceback; traceback.print_exc()
        return {}


# ===== 写 Excel =====
def write_excel(all_prices: dict) -> int:
    from changelog import record_changes, SOURCE_AUTO_CRON, norm_value, current_commit_sha
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    row = get_row(ws)
    written, skipped = [], []
    diffs = []

    # 日期行标识（用于变更记录）
    dv = ws.cell(row, 1).value
    if isinstance(dv, datetime.datetime):
        date_str = dv.strftime("%Y-%m-%d")
    elif isinstance(dv, str):
        date_str = dv.strip()
    else:
        date_str = TODAY

    for col, (code, src) in COL_MAP.items():
        old_val = norm_value(ws.cell(row, col).value)
        if code in all_prices and all_prices[code] is not None:
            new_val = float(all_prices[code])
            ws.cell(row, col).value = new_val
            written.append(code)
            if old_val != new_val:
                diffs.append({"date_row": date_str, "code": code, "old": old_val, "new": new_val})
        else:
            skipped.append(code)
    wb.save(EXCEL_PATH)
    print(f"  行 {row}: 写入 {len(written)}/{len(COL_MAP)} 项")
    if skipped:
        print(f"  缺: {', '.join(skipped)}")

    # 变更留痕（自动抓取覆盖）
    if diffs:
        sha = current_commit_sha(str(SCRIPT_DIR))
        n = record_changes(diffs, source=SOURCE_AUTO_CRON, editor="—", commit=sha)
        print(f"  📝 变更留痕 {n} 条（auto_cron）")
    return row


# ===== 导出 + Push =====
def run_git(args, timeout=60, retries=3):
    """带超时与重试的 git 调用，规避沙箱网络 github.com 偶发挂起。"""
    last = None
    for i in range(retries):
        try:
            r = subprocess.run(["git"] + args, cwd=str(SCRIPT_DIR),
                               capture_output=True, text=True, timeout=timeout)
            if r.returncode == 0:
                return r
            msg = (r.stdout + r.stderr).lower()
            if any(k in msg for k in ("timed out", "timeout", "could not resolve",
                                      "connection", "403", "502", "503", "non-fast-forward")):
                print(f"  git {' '.join(args)} 可重试({msg.strip()[:100]}) 重试 {i+1}/{retries}")
                time.sleep(3); last = r; continue
            return r
        except subprocess.TimeoutExpired:
            print(f"  git {' '.join(args)} 超时({timeout}s) 重试 {i+1}/{retries}")
            last = None; time.sleep(3)
    return last


def export_and_push(row: int):
    print("  导出 data.json...")
    subprocess.run([sys.executable, str(SCRIPT_DIR / "export_excel_to_json.py")],
        cwd=str(SCRIPT_DIR), capture_output=True, text=True)
    r = run_git(["add", "docs/data.json"])
    if r.returncode != 0:
        print(f"  ⚠️ git add: {(r.stdout + r.stderr).strip()}")
        return False
    r = run_git(["commit", "-m", f"全品种更新 {TODAY} row={row}"])
    if r.returncode != 0 and "nothing to commit" not in (r.stdout + r.stderr).lower():
        print(f"  ⚠️ git commit: {(r.stdout + r.stderr).strip()}")
        return False
    r = run_git(["push"], timeout=90, retries=4)
    if r is None or (r.returncode != 0 and "nothing to commit" not in (r.stdout + r.stderr).lower()):
        print(f"  ⚠️ git push 失败: {((r.stdout + r.stderr).strip() if r else 'timeout')}")
        return False
    print("  ✅ 推送成功")
    return True


async def main():
    print(f"\n{'='*50}")
    print(f"  全品种抓取 · {TODAY}")
    print(f"{'='*50}\n")

    # 第一步: git pull + 同步在线编辑到 Excel
    print("[0/4] 同步在线编辑...")
    try:
        run_git(["pull"], timeout=60, retries=2)
        print("  git pull 完成")
        # 将 data.json 中的在线编辑回写到 Excel
        from sync_from_web import sync_to_excel
        sync_to_excel()
    except Exception as e:
        print(f"  ⚠️ 同步失败: {e}（继续执行）")

    all_prices = {}

    all_prices.update(await fetch_ccmn())
    print()
    all_prices.update(await fetch_smm())
    print()
    # 亚洲金属网闻喜镁锭覆盖 SMM 的 Wenxi_MG（项目要求主源为亚洲金属网）
    all_prices.update(await fetch_asianmetal())
    print()
    tungsten = fetch_tungsten()
    if tungsten: all_prices.update(tungsten)
    wti = fetch_wti()
    if wti: all_prices.update(wti)
    steel = await fetch_steel()
    if steel: all_prices.update(steel)
    sci99 = await fetch_sci99()
    if sci99: all_prices.update(sci99)
    print()

    print(f"[写表] ", end="")
    row = write_excel(all_prices)

    print(f"[发布] ", end="")
    export_and_push(row)

    print(f"\n  已更新 {len(all_prices)}/{len(COL_MAP)} 品种")
    print(f"  看板: https://lekelin2046.github.io/jinggong-commodity-monitor/\n")


if __name__ == "__main__":
    asyncio.run(main())
