"""钨粉晚点补查脚本（21:00 跑）

- 只查钨粉一项（轻量，10-15秒）
- 如果今天的钨粉行还是空（标黄），用最新数据补上
- 其他列不碰
- 跑完发简短消息给主人
"""
import asyncio
import logging
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

sys.path.insert(0, str(Path(__file__).parent))
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s")
logger = logging.getLogger(__name__)

EXCEL_PATH = Path("/Users/siqi/Desktop/AI/jinggong-commodity-monitor/2026年有色金属市场价格共享(2).xlsx")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
SCREENSHOTS_ROOT = Path("/Users/siqi/Desktop/AI/jinggong-commodity-monitor/screenshots")
TODAY_STR = datetime.now().strftime("%Y-%m-%d")


async def retry_tungsten():
    """晚点补查钨粉"""
    from fill_and_verify import get_tungsten_price
    
    print(f"[{datetime.now().strftime('%H:%M')}] 钨粉晚点补查 ({TODAY_STR})")
    
    # 1. 先检查 Excel 今天那行钨粉是否已填
    wb = openpyxl.load_workbook(str(EXCEL_PATH))
    ws = wb["日均价（中钨在线 原油）"] if "日均价（中钨在线 原油）" in wb.sheetnames else wb.create_sheet("日均价（中钨在线 原油）")
    target_row = None
    for r in range(3, ws.max_row + 2):
        v = ws.cell(row=r, column=1).value
        if v and isinstance(v, datetime) and v.strftime("%Y-%m-%d") == TODAY_STR:
            target_row = r
            break
    
    if target_row is None:
        # 今天的行不存在，先建一个
        target_row = ws.max_row + 1
        ws.cell(row=target_row, column=1, value=datetime.strptime(TODAY_STR, "%Y-%m-%d"))
        ws.cell(row=target_row, column=2, value="钨粉")
        print(f"  今天行不存在，新建 Row{target_row}")
    
    # 检查 Col C 钨粉含税价是否已有
    existing_price = ws.cell(row=target_row, column=3).value
    if existing_price and existing_price > 0:
        print(f"  ✅ 钨粉已有价 ({existing_price})，跳过")
        wb.close()
        return
    
    # 2. 查钨粉
    print("  查钨粉...")
    tungsten = await get_tungsten_price() if asyncio.iscoroutinefunction(get_tungsten_price) else None
    
    if not tungsten:
        # get_tungsten_price 不是 async，包一下
        tungsten = get_tungsten_price()
    
    if tungsten and tungsten > 0:
        ws.cell(row=target_row, column=3, value=tungsten)
        # 清掉黄填充
        ws.cell(row=target_row, column=3).fill = PatternFill(fill_type=None)
        print(f"  ✅ 钨粉已补: {tungsten} 元/千克")
        wb.save(str(EXCEL_PATH))
        print(f"  Excel 已保存")
    else:
        print(f"  ⚠️ 钨粉仍未出，继续标黄")
    
    wb.close()


if __name__ == "__main__":
    # 直接同步调（get_tungsten_price 是 sync 的）
    from fill_and_verify import get_tungsten_price
    
    print(f"[{datetime.now().strftime('%H:%M')}] 钨粉晚点补查 ({TODAY_STR})")
    
    wb = openpyxl.load_workbook(str(EXCEL_PATH))
    ws = wb["日均价（中钨在线 原油）"] if "日均价（中钨在线 原油）" in wb.sheetnames else wb.create_sheet("日均价（中钨在线 原油）")
    
    # 找今天的行
    target_row = None
    for r in range(3, ws.max_row + 2):
        v = ws.cell(row=r, column=1).value
        if v and isinstance(v, datetime) and v.strftime("%Y-%m-%d") == TODAY_STR:
            target_row = r
            break
    
    if target_row is None:
        target_row = ws.max_row + 1
        ws.cell(row=target_row, column=1, value=datetime.strptime(TODAY_STR, "%Y-%m-%d"))
        ws.cell(row=target_row, column=2, value="钨粉")
        print(f"  今天行不存在，新建 Row{target_row}")
    
    existing_price = ws.cell(row=target_row, column=3).value
    if existing_price and isinstance(existing_price, (int, float)) and existing_price > 0:
        print(f"  ✅ 钨粉已有价 ({existing_price})，跳过")
        wb.close()
        sys.exit(0)
    
    print("  查钨粉...")
    tungsten = get_tungsten_price()
    
    if tungsten and tungsten > 0:
        cell = ws.cell(row=target_row, column=3)
        cell.value = tungsten
        cell.fill = PatternFill(fill_type=None)
        # 同步更新 D 公式
        ws.cell(row=target_row, column=4, value=f"=IFERROR(C{target_row}/1.13,\"\")")
        wb.save(str(EXCEL_PATH))
        print(f"  ✅ 钨粉已补: {tungsten} 元/千克（已写入 Excel）")
        sys.exit(0)
    else:
        print(f"  ⚠️ 钨粉仍未出，保持标黄")
        wb.close()
        sys.exit(1)
