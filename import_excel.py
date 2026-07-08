"""一次性脚本：从 Excel 导入全部历史数据到 SQLite
python3 import_excel.py
"""
import sqlite3, sys
from pathlib import Path
from datetime import datetime
import openpyxl

PROJECT = Path(__file__).parent
EXCEL_PATH = PROJECT / "2026年有色金属市场价格.xlsx"
DB_PATH = PROJECT / "data" / "prices.db"

# Sheet1 列 → 品种代码（参照前面摸清的映射）
VARIETY_MAP = {
    2:  "ADC12",
    3:  "A380",
    4:  "AlSi9Cu3",
    5:  "A356",
    6:  "A00_AL",
    7:  "CU",
    8:  "SI_441",       # 硅441中间价
    9:  "SI_3303",      # 硅3303中间价
    10: "MG",
    11: "MN",
    12: "SI_331",       # 硅331中间价
    13: "Wenxi_MG",
    14: "AM60B",
    15: "AZ91D",
    16: "W",
    # 17 是 WTI（在 Sheet1 的 Col17）
}

# 品种中文名（供看板展示）
VARIETY_NAMES = {
    "ADC12":      "ADC12",
    "A380":       "A380",
    "AlSi9Cu3":   "AlSi9Cu3",
    "A356":       "A356",
    "A00_AL":     "A00铝",
    "CU":         "铜",
    "SI_441":     "硅441",
    "SI_3303":    "硅3303",
    "MG":         "镁",
    "MN":         "电解锰",
    "SI_331":     "硅331",
    "Wenxi_MG":   "闻喜镁锭",
    "AM60B":      "AM60B",
    "AZ91D":      "AZ91D",
    "W":          "钨粉",
    "WTI":        "WTI原油",
}

# 品种展示顺序
VARIETY_ORDER = [
    "ADC12", "A380", "AlSi9Cu3", "A356",
    "A00_AL", "CU",
    "SI_441", "SI_3303", "SI_331",
    "MG", "MN",
    "Wenxi_MG", "AM60B", "AZ91D",
    "W", "WTI",
]


def create_db():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("""
        CREATE TABLE IF NOT EXISTS prices (
            date     TEXT NOT NULL,
            code     TEXT NOT NULL,
            price    REAL,
            source   TEXT DEFAULT 'excel',
            PRIMARY KEY (date, code)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS variety_names (
            code TEXT PRIMARY KEY,
            name TEXT NOT NULL
        )
    """)
    return conn


def import_sheet1(conn, ws):
    """导入 Sheet1 的 15 个品种 × 所有日期行"""
    inserted = 0
    skipped_empty = 0

    for r in range(3, ws.max_row + 1):
        date_val = ws.cell(row=r, column=1).value
        if not date_val:
            continue

        try:
            if isinstance(date_val, datetime):
                date_str = date_val.strftime("%Y-%m-%d")
            else:
                date_str = str(date_val)[:10]
            datetime.strptime(date_str, "%Y-%m-%d")  # 校验
        except Exception:
            continue

        for col, code in VARIETY_MAP.items():
            val = ws.cell(row=r, column=col).value
            if val is None:
                # None → "——" 跳过不存，看板显示横杠
                continue

            if isinstance(val, str):
                val = val.replace(",", "").replace("，", "").strip()
                if not val or val in ("——", "—", "-", "—", "N/A", ""):
                    continue
                try:
                    price = float(val)
                except Exception:
                    continue
            elif isinstance(val, (int, float)):
                price = float(val)
            else:
                continue

            # 过滤离谱值
            if price <= 0 or price > 500000:
                continue

            try:
                conn.execute(
                    "INSERT OR REPLACE INTO prices (date, code, price, source) VALUES (?,?,?,?)",
                    (date_str, code, price, "excel")
                )
                inserted += 1
            except Exception as e:
                print(f"  ⚠️ {date_str} {code}={price}: {e}")

    print(f"  Sheet1: 写入 {inserted} 条，跳过空值 {skipped_empty} 行")


def import_sheet2(conn, ws2):
    """导入 Sheet2 的钨粉 + WTI（格式较乱，日期在 ColA 和 ColF）"""
    inserted = 0

    for r in range(2, ws2.max_row + 1):
        # 钨粉：ColA=日期, ColC=价格
        date_a = ws2.cell(row=r, column=1).value
        name_b = ws2.cell(row=r, column=2).value
        price_c = ws2.cell(row=r, column=3).value

        if date_a and isinstance(date_a, datetime) and name_b and "钨粉" in str(name_b) and price_c:
            try:
                p = float(price_c)
                if 100 < p < 5000:
                    dt = date_a.strftime("%Y-%m-%d")
                    conn.execute(
                        "INSERT OR REPLACE INTO prices (date, code, price, source) VALUES (?,?,?,?)",
                        (dt, "W", p, "excel_sheet2")
                    )
                    inserted += 1
            except Exception:
                pass

        # WTI：ColE="WTI原油", ColF=日期, ColG=价格
        name_e = ws2.cell(row=r, column=5).value
        date_f = ws2.cell(row=r, column=6).value
        price_g = ws2.cell(row=r, column=7).value

        if name_e and "WTI" in str(name_e) and date_f and isinstance(date_f, datetime) and price_g:
            try:
                p = float(price_g)
                if 1 < p < 200:
                    dt = date_f.strftime("%Y-%m-%d")
                    conn.execute(
                        "INSERT OR REPLACE INTO prices (date, code, price, source) VALUES (?,?,?,?)",
                        (dt, "WTI", p, "excel_sheet2")
                    )
                    inserted += 1
            except Exception:
                pass

    print(f"  Sheet2: 写入 {inserted} 条")


def insert_variety_names(conn):
    for code, name in VARIETY_NAMES.items():
        conn.execute(
            "INSERT OR REPLACE INTO variety_names (code, name) VALUES (?,?)",
            (code, name)
        )


def main():
    print("=" * 50)
    print("📥 导入 Excel 历史数据 → SQLite")
    print(f"  Excel: {EXCEL_PATH}")
    print(f"  DB:    {DB_PATH}")
    print()

    if not EXCEL_PATH.exists():
        print(f"❌ Excel 文件不存在!")
        sys.exit(1)

    conn = create_db()

    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True)
    print("📄 加载 Excel ...")

    ws1 = wb["日均价（2026年市场）"]
    import_sheet1(conn, ws1)

    ws2 = wb["日均价（中钨在线 原油）"]
    import_sheet2(conn, ws2)

    wb.close()

    insert_variety_names(conn)

    # 统计
    total = conn.execute("SELECT COUNT(*) FROM prices").fetchone()[0]
    dates = conn.execute("SELECT COUNT(DISTINCT date) FROM prices").fetchone()[0]
    codes = conn.execute("SELECT COUNT(DISTINCT code) FROM prices").fetchone()[0]
    print(f"\n✅ 导入完成: {total} 条记录, {dates} 天, {codes} 个品种")

    # 看各品种数据量
    print("\n品种数据量:")
    for row in conn.execute(
        "SELECT code, COUNT(*) as cnt, MIN(date), MAX(date) FROM prices GROUP BY code ORDER BY code"
    ):
        print(f"  {row[0]:12}  {row[1]:4} 条  {row[2]} ~ {row[3]}")

    conn.commit()
    conn.close()


if __name__ == "__main__":
    main()
