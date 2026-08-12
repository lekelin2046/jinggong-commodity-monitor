#!/usr/bin/env python3
"""
5PM 补抓检查 — 检查今日全品种（25项）是否填全，缺则补抓

用法: python3 daily_check_missed.py
"""

import sys, datetime
from pathlib import Path

SCRIPT_DIR = Path(__file__).parent
sys.path.insert(0, str(SCRIPT_DIR))

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed", file=sys.stderr)
    sys.exit(1)

EXCEL_PATH = SCRIPT_DIR / "2026年有色金属市场价格.xlsx"
SHEET_NAME = "日均价（2026年市场）"
SMM_COLS = [2,3,4,5,14,15,13]  # SMM: 2-5, 13-15
CCMN_COLS = [6,7,8,9,10,11,12]  # CCMN: 6-12
TOTAL = 25  # 全品种数（列2-26）
ALL_COLS = list(range(2, 27))  # 2-26（覆盖全部 25 品种）

today = datetime.date.today()
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb[SHEET_NAME]

# 找今天行
target_row = None
for r in range(2, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if isinstance(v, datetime.datetime) and v.date() == today:
        target_row = r
        break

if not target_row:
    print(f"→ 今天 ({today}) 无数据行，触发全品种补抓")
    import subprocess
    subprocess.run([sys.executable, str(SCRIPT_DIR / "daily_update_all.py")], cwd=str(SCRIPT_DIR))
    sys.exit()

# 检查
missing = [c for c in ALL_COLS if ws.cell(target_row, c).value is None]
if not missing:
    print(f"✓ 今日 #{target_row} 行 {TOTAL}/{TOTAL} 品种已全部填全，无需补抓")
    sys.exit(0)

missing_labels = {2:"ADC12",3:"A380",4:"AlSi9Cu3",5:"A356",
    6:"A00_AL",7:"CU",8:"SI_441",9:"SI_3303",10:"MG",11:"MN",
    12:"SI_331",13:"Wenxi_MG",14:"AM60B",15:"AZ91D",16:"W",17:"WTI",
    18:"IRON_ORE",19:"COKE",20:"SS_304",21:"SS_409",22:"SS_439",
    23:"SS_441",24:"NICKEL_IRON",25:"HIGH_CARBON_FECR",26:"ADC12_JAPAN_CIF"}

missing_names = [missing_labels.get(c, f"col{c}") for c in missing]
print(f"→ 今日 #{target_row} 行缺 {len(missing)}/{TOTAL} 品种: {', '.join(missing_names)}，触发补抓")
import subprocess
subprocess.run([sys.executable, str(SCRIPT_DIR / "daily_update_all.py")], cwd=str(SCRIPT_DIR))
