"""Excel 填充器 v2

修正表匹配逻辑：
- 精工16品种 vs 有色金属共享表 → 只有非铁金属重叠
- 匹配: ADC12(Col2), A00铝(Col6), 铜(Col7)
- 注意: Col8"金属硅441" ≠ 不锈钢441！这是工业硅，别填错！
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict

import openpyxl

from jinggong_monitor import get_varieties

logger = logging.getLogger("jinggong.excel_filler")

_EXCEL_PATH = Path.home() / "Desktop" / "工作" / "招聘" / "2026年有色金属市场价格共享.xlsx"

# 精工品种 → Excel 列匹配（精确匹配）
_VARIETY_TO_EXCEL_COL = {
    # 格式: variety_id → (column_index, match_keywords)
    "ADC12":   (2, ["ADC12", "ADC 12"]),
    "A00_AL":  (6, ["A00铝", "A00 铝", "铝中间价"]),
    "CU":      (7, ["铜中间价", "铜"]),
}


def _find_columns(ws) -> dict:
    """从表头行自动匹配列"""
    col_map = {}
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        v = str(ws.cell(row=1, column=col_idx).value or "").strip()
        if v:
            headers[col_idx] = v

    for vid, (default_col, keywords) in _VARIETY_TO_EXCEL_COL.items():
        # 先精确匹配表头
        for col_idx, header in headers.items():
            if any(kw in header for kw in keywords):
                col_map[vid] = col_idx
                break
        # 没匹配到用默认列
        if vid not in col_map:
            col_map[vid] = default_col

    logger.info("列匹配: %s", {vid: headers.get(c, f"Col{c}") for vid, c in col_map.items()})
    return col_map


def fill_prices(
    prices: dict[str, float],
    target_date: str = None,
) -> str:
    """将价格写入共享 Excel 对应列

    Args:
        prices: {variety_id: price}
        target_date: 日期 YYYY-MM-DD

    Returns:
        文件路径
    """
    path = _EXCEL_PATH
    if not path.exists():
        logger.error("Excel 不存在: %s", path)
        return ""

    today = target_date or datetime.now().strftime("%Y-%m-%d")

    wb = openpyxl.load_workbook(str(path))
    ws = wb.active

    col_map = _find_columns(ws)

    # 找日期列
    date_col = 1  # Col 1 = 日期
    existing_row = None
    for row_idx in range(2, ws.max_row + 1):
        cell_val = str(ws.cell(row=row_idx, column=date_col).value or "")
        if today in cell_val:
            existing_row = row_idx
            break

    if existing_row is None:
        existing_row = ws.max_row + 1
        ws.cell(row=existing_row, column=date_col, value=today)

    # 填入价格
    filled = {}
    for vid, col_idx in col_map.items():
        price = prices.get(vid)
        if price is not None and price > 0:
            ws.cell(row=existing_row, column=col_idx, value=round(price, 2))
            filled[vid] = price
            logger.info("→ Col%d %s: %.2f", col_idx, vid, price)

    wb.save(str(path))
    wb.close()

    logger.info("填表完成: %s 第%d行, %d/3列", today, existing_row, len(filled))
    return str(path)


def fill_excel(
    prices: dict[str, float],
    excel_path: Optional[str] = None,
    target_date: Optional[str] = None,
) -> str:
    """兼容旧接口"""
    return fill_prices(prices, target_date)
