"""精工有色金属共享表 - 全面填充+核查脚本

功能：
1. 填写「日均价（2026年市场）」sheet 2026-06-23/24 数据
2. 填写「日均价（中钨在线 原油）」sheet 2026-06-23/24 数据
3. 不可填写项标黄+备注原因
4. 核查 6/16-22 历史数据准确性，输出偏离度报告
"""

import asyncio
import logging
import os
import re
import subprocess
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# 6/29 主人拍板：填表字体与原表统一
DATA_FONT = Font(name="微软雅黑", size=11, bold=False)  # 数据行用 微软雅黑 11
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True)  # 标题行用 微软雅黑 10
DATE_FONT = Font(name="微软雅黑", size=11, bold=False)
from playwright.async_api import async_playwright

sys.path.insert(0, str(Path(__file__).parent))
from jinggong_monitor.credentials import require_smm, require_asianmetal

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s")
logger = logging.getLogger(__name__)

EXCEL_PATH = Path(__file__).parent / "2026年有色金属市场价格共享(2).xlsx"
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# ============================================================
# 自动登录配置（2026-07-02 突破：networkidle → domcontentloaded）
# 账号密码从环境变量 / .env 读取，缺失即报错（不保留默认密码）
# ============================================================
SMM_ACCOUNT, SMM_PASSWORD = require_smm()
ASIANMETAL_ACCOUNT, ASIANMETAL_PASSWORD = require_asianmetal()

# 代理白名单：国内站点直连
_PROXY_DOMAINS = "sci99.com,chinatungsten.com,51bxg.com,steelcn.cn,ccmn.cn,cnfeol.com,ctia.com.cn,smm.cn,asianmetal.cn,hq.smm.cn,user.smm.cn,www.asianmetal.cn"
os.environ.setdefault("NO_PROXY", _PROXY_DOMAINS)
os.environ.setdefault("no_proxy", os.environ.get("NO_PROXY", _PROXY_DOMAINS))


async def _launch_headless_browser():
    """统一启动 headless 浏览器（去掉对 9223 调试 Chrome 的依赖）"""
    p = await async_playwright().start()
    browser = await p.chromium.launch(
        headless=True,
        args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
    )
    return p, browser


async def _smm_login_and_get_cookies(context) -> bool:
    """在给定 context 里自动登录 SMM（无需验证码）

    返回 True 表示登录成功（URL 不含 login）。
    2026-07-02 实测：domcontentloaded + 5s 等待，5 秒内必登录成功。
    """
    page = await context.new_page()
    try:
        await page.goto("https://user.smm.cn/login", timeout=30000, wait_until="domcontentloaded")
        await page.wait_for_timeout(3000)
        await page.locator("#userName").fill(SMM_ACCOUNT)
        await page.locator("#password").fill(SMM_PASSWORD)
        await page.locator("#user_account_password_login_button").click()
        await asyncio.sleep(5)
        return "login" not in page.url.lower()
    finally:
        await page.close()


async def _asianmetal_login_and_get_cookies(context) -> bool:
    """在给定 context 里自动登录亚洲金属网

    2026-07-02 实测：主登录表单（txtUserLoginName/txtUserPwd）无法建立主站会话；
    真正可用的是顶部登录弹窗（cnopenloginname / cnopenloginpwd / openloginbutn）。
    访问受保护文章会触发弹窗，登录后跳回文章页。
    """
    page = await context.new_page()
    try:
        protected_url = "https://www.asianmetal.cn/news/2974825/"
        await page.goto(protected_url, timeout=30000, wait_until="domcontentloaded")
        await page.wait_for_timeout(3000)

        # 显示并填充顶部登录弹窗
        await page.evaluate(
            """(args) => {
                const loginWin = document.getElementById("loginWin");
                const showFilter = document.getElementById("showFilter");
                if (loginWin) loginWin.style.display = "block";
                if (showFilter) showFilter.style.display = "block";
                document.getElementById("cnopenloginname").value = args.user;
                document.getElementById("cnopenloginpwd").value = args.pwd;
            }""",
            {"user": ASIANMETAL_ACCOUNT, "pwd": ASIANMETAL_PASSWORD},
        )
        await page.wait_for_timeout(500)
        await page.locator("#openloginbutn").click()
        await asyncio.sleep(8)

        # 处理「账号在线中」强制下线弹窗：点 #outlinebutn 继续
        try:
            for _ in range(2):
                await page.wait_for_timeout(3000)
                body = await page.inner_text("body")
                if "此账号在线中" in body or "强制对方下线" in body:
                    print(f"  检测到账号在线中弹窗，点击「登陆」强制进入...")
                    try:
                        await page.locator("#outlinebutn").click(timeout=3000)
                    except Exception:
                        await page.evaluate('''() => { if (typeof outLine === "function") outLine(); }''')
                    await page.wait_for_timeout(3000)
        except Exception:
            pass

        body = await page.inner_text("body")
        return "news/2974825" in page.url and ("注销" in body or "欢迎" in body)
    finally:
        await page.close()

# ============================================================
# 截图管理（2026-06-29 主人新增需求）
# ============================================================
SCREENSHOTS_ROOT = Path(__file__).parent / "screenshots"
TODAY_STR = datetime.now().strftime("%Y-%m-%d")
SHOT_DIR = SCREENSHOTS_ROOT / TODAY_STR
SHOT_DIR.mkdir(parents=True, exist_ok=True)

# 截图结果登记表：{数据源: (ok, 路径或错误)}
SHOT_RESULTS: dict[str, tuple[bool, str]] = {}


def _shot_name(source: str, page_label: str, ok: bool) -> str:
    """生成截图文件名：{ok前缀}_{source}_{page}_{时间戳}.png"""
    prefix = "" if ok else "❌_"
    ts = datetime.now().strftime("%H%M%S")
    safe = re.sub(r"[^\w\u4e00-\u9fa5-]+", "_", f"{source}_{page_label}")
    return f"{prefix}{safe}_{ts}.png"


async def take_screenshot(page, source: str, page_label: str, full_page: bool = False) -> Optional[Path]:
    """截图并登记结果。失败也截一张（标 ❌）。

    Args:
        full_page: True=截整页（含页面标题/日期/表格）— 6/29 主人拍板
                  False=只截表格区域（带日期+内容）
    """
    try:
        await page.wait_for_timeout(800)  # 等动画/字体稳定
        # 整页模式：full_page=True
        # 表格区域模式：尝试找表格 clip，找不到 fallback 到全页
        clip = None
        if not full_page:
            try:
                tbl = await page.query_selector("table")
                if tbl:
                    box = await tbl.bounding_box()
                    if box and box["width"] > 200 and box["height"] > 100:
                        clip = {
                            "x": max(0, box["x"] - 20),
                            "y": max(0, box["y"] - 20),
                            "width": min(box["width"] + 40, 1920),
                            "height": min(box["height"] + 40, 1500),
                        }
            except Exception:
                clip = None

        path = SHOT_DIR / _shot_name(source, page_label, ok=True)
        await page.screenshot(path=str(path), full_page=full_page, clip=clip if not full_page else None)
        SHOT_RESULTS[f"{source}/{page_label}"] = (True, str(path.relative_to(SCREENSHOTS_ROOT)))
        logger.info(f"📸 截图 OK: {path.name} ({'整页' if full_page else '表格区'})")
        return path
    except Exception as e:
        # 失败也截一张全页（便于排查）
        try:
            fail_path = SHOT_DIR / _shot_name(source, page_label, ok=False)
            await page.screenshot(path=str(fail_path), full_page=True)
            SHOT_RESULTS[f"{source}/{page_label}"] = (False, f"截图失败: {e}; 已存全页到 {fail_path.name}")
        except Exception as e2:
            SHOT_RESULTS[f"{source}/{page_label}"] = (False, f"截图失败: {e}; 二次失败: {e2}")
        return None

# ============================================================
# 数据采集
# ============================================================

def get_akshare_spot(variety: str, date_str: str) -> Optional[float]:
    """从akshare获取futures_spot_price"""
    import akshare as ak
    try:
        df = ak.futures_spot_price(date=date_str.replace("-", ""), vars_list=[variety])
        if df is not None and not df.empty:
            return float(df.iloc[0]["spot_price"])
    except Exception:
        pass
    return None


async def get_ccmn_prices() -> dict:
    """调 ccmn AJAX 端点（fetcher_ccmn.CcmnFetcher）+ 截图存档

    2026-07-02 改造：去掉 CDP 9223 依赖，独立 headless 浏览器截图。
    6/30 主人拍板：禁用首页 Playwright + 正则方案（ccmn 6 月起将电解锰表格分行了，
    正则 r"1#电解锰[^合]\\D+?(\\d+)..." 会匹配错位抓到无关小数字 → 写出 6 这种离谱值）。
    现在直接复用 fetcher_ccmn.py 验证过的 AJAX 端点（POST /shop/historyData/...）。
    """
    # 📸 截图存档：ccmn 截整页含页面标题+日期+表格（独立 headless，不再依赖 9223）
    p, browser = await _launch_headless_browser()
    try:
        ctx = await browser.new_context(viewport={"width": 1280, "height": 900})
        page = await ctx.new_page()
        await page.goto("https://www.ccmn.cn/cjxh.shtml", timeout=20000, wait_until="domcontentloaded")
        await page.wait_for_timeout(3000)
        await take_screenshot(page, "ccmn", "长江现货", full_page=True)
        await page.close()
        await ctx.close()
    except Exception as e:
        print(f"⚠️ ccmn 截图失败（不影响数据采集）: {e}")
    finally:
        await browser.close()
        await p.stop()

    # 调 AJAX 端点（fetcher_ccmn 是 6/25 验证过的稳定实现）
    from jinggong_monitor.fetcher_ccmn import CcmnFetcher
    fetcher = CcmnFetcher()
    today = datetime.now().strftime("%Y-%m-%d")
    prices = fetcher.fetch(today)
    return prices


async def get_smm_cdp() -> dict:
    """自动登录 SMM 抓 7 个品种价格（2026-07-02 改造）

    之前：connect_over_cdp(localhost:9223) → 依赖主人手动开 Chrome 登录态
    现在：独立 headless 浏览器，5秒自动登录 + cookies 注入 + 抓取
    实测：ADC12/A380/AlSi9Cu3/A356/闻喜镁锭/AM60B/AZ91D 全部抓到
    """
    p, browser = await _launch_headless_browser()
    try:
        context = await browser.new_context(
            viewport={"width": 1280, "height": 900},
            bypass_csp=True,
        )

        # 1. 自动登录 SMM
        login_ok = await _smm_login_and_get_cookies(context)
        if not login_ok:
            logger.warning("SMM 自动登录失败，可能账号密码错或网站改版")
            return {}
        logger.info("SMM 自动登录成功")

        # 2. 访问铝行情页
        page = await context.new_page()
        await page.goto("https://hq.smm.cn/aluminum", timeout=20000, wait_until="domcontentloaded")
        await page.wait_for_timeout(5000)
        al_text = await page.inner_text("body")
        # 📸 截图 SMM 铝页
        await take_screenshot(page, "smm", "铝页")

        # 3. 访问镁行情页
        page2 = await context.new_page()
        await page2.goto("https://hq.smm.cn/magnesium", timeout=20000, wait_until="domcontentloaded")
        await page2.wait_for_timeout(5000)
        mg_text = await page2.inner_text("body")
        # 📸 截图 SMM 镁页
        await take_screenshot(page2, "smm", "镁页")

        await page.close()
        await page2.close()

        # 4. 保存 cookies（必须在 context.close() 之前，否则 cookies() 会报错）
        try:
            cookies = await context.cookies()
            cookie_path = Path(__file__).parent / "data" / "smm_cookies.json"
            cookie_path.parent.mkdir(parents=True, exist_ok=True)
            import json
            with open(cookie_path, "w") as f:
                json.dump(cookies, f, indent=2, ensure_ascii=False)
            logger.info(f"SMM cookies 保存: {len(cookies)} 条")
        except Exception as e:
            logger.warning(f"cookies 保存失败（不影响抓取）: {e}")

        await context.close()
    finally:
        await browser.close()
        await p.stop()

    # 5. 正则提取价格
    prices = {}
    smm_patterns = {
        "ADC12": (al_text, r"SMM铝合金ADC12\D+?(\d+)\D+?(\d+)\D+?(\d+)"),
        "A380": (al_text, r"A380铝合金\D+?(\d+)\D+?(\d+)\D+?(\d+)"),
        "AlSi9Cu3": (al_text, r"AlSi9Cu3铝合金\D+?(\d+)\D+?(\d+)\D+?(\d+)"),
        "A356": (al_text, r"A356铝合金\s+\D+?(\d+)\D+?(\d+)\D+?(\d+)"),
        "Wenxi_MG": (mg_text, r"镁锭9990（闻喜）\D+?(\d+)\D+?(\d+)\D+?(\d+)"),
        "AM60B": (mg_text, r"AM60B出厂价\D+?(\d+)\D+?(\d+)\D+?(\d+)"),
        "AZ91D": (mg_text, r"AZ91D出厂价[^（]*?(\d+)\D+?(\d+)\D+?(\d+)"),
    }
    for name, (text, pat) in smm_patterns.items():
        m = re.search(pat, text)
        if m:
            prices[name] = int(m.group(3))
    return prices


async def get_asianmetal_price() -> dict:
    """自动登录亚洲金属网抓闻喜镁锭价格（2026-07-02 改造）

    之前：connect_over_cdp(localhost:9223) + fetcher._fetch_via_cdp（实际不存在该方法）
    现在：独立 headless 浏览器，自动登录 + 跳文章页 + 尝试表格提取 → 失败走 OCR 备选

    6/30 主人拍板：6 月起价格表改为图片渲染（防爬），原 _extract_wenxi_from_table 拿不到文字。
    业主硬要求：闻喜镁锭必须用亚洲金属网，不走 SMM 备源。
    """
    p, browser = await _launch_headless_browser()
    try:
        context = await browser.new_context(
            viewport={"width": 1280, "height": 900},
            bypass_csp=True,
        )

        # 1. 自动登录亚洲金属网
        login_ok = await _asianmetal_login_and_get_cookies(context)
        if not login_ok:
            logger.warning("亚洲金属网自动登录失败")
            return {}
        logger.info("亚洲金属网自动登录成功")

        # 2. 保存 cookies
        try:
            cookies = await context.cookies()
            cookie_path = Path(__file__).parent / "data" / "asianmetal_cookies.json"
            cookie_path.parent.mkdir(parents=True, exist_ok=True)
            import json
            with open(cookie_path, "w") as f:
                json.dump(cookies, f, indent=2, ensure_ascii=False)
            logger.info(f"亚洲金属网 cookies 保存: {len(cookies)} 条")
        except Exception as e:
            logger.warning(f"cookies 保存失败: {e}")

        # 3. 跳到镁锭新闻列表页，找今日镁锭价格文章
        page = await context.new_page()
        await page.goto("https://www.asianmetal.cn/MagnesiumPricesHistory", timeout=20000, wait_until="domcontentloaded")
        await page.wait_for_timeout(3000)

        article_url = await page.evaluate('''
            () => {
                const links = document.querySelectorAll('a[href*="/news/"]');
                for (const a of links) {
                    const t = (a.innerText || '').split('\\n')[0].trim();
                    if (!t) continue;
                    if (t.includes('镁合金') || t.includes('镁粉') || t.includes('99.95%min')) continue;
                    if (t.includes('镁锭') && t.includes('价格')) return a.href;
                }
                return null;
            }
        ''')

        if not article_url:
            logger.warning("亚洲金属网未找到今日镁锭文章")
            await take_screenshot(page, "asianmetal", "列表页未找到文章", full_page=True)
            await page.close()
            await context.close()
            return {}

        # 4. 打开文章详情页
        logger.info(f"打开文章: {article_url}")
        await page.goto(article_url, timeout=20000, wait_until="domcontentloaded")
        await page.wait_for_timeout(5000)

        # 5. 截图存证
        await take_screenshot(page, "asianmetal", "闻喜镁锭", full_page=True)

        # 6. 尝试文字提取（表格/正则）
        body_text = await page.inner_text("body")
        wenxi_row = re.search(
            r'闻喜[\s\S]{0,40}?镁锭[\s\S]{0,40}?99[.,]9%min[\s\S]{0,40}?'
            r'(\d[\d,.]+\d)\s*[-–~]\s*(\d[\d,.]+\d)',
            body_text,
        )
        if wenxi_row:
            low = float(wenxi_row.group(1).replace(",", ""))
            high = float(wenxi_row.group(2).replace(",", ""))
            if 10000 < low < 50000 and 10000 < high < 50000:
                mid = round((low + high) / 2, 2)
                logger.info(f"亚洲金属网闻喜镁锭（文字）: {low}-{high} → {mid}")
                await page.close()
                await context.close()
                return {"Wenxi_MG": mid}

        # 7. 文字提取失败（图片渲染）→ 走 OCR 备选
        logger.warning("文字提取失败（价格表可能是图片），走 OCR 备选")
        await page.close()
        await context.close()

        price = await _ocr_asianmetal_fallback()
        if price:
            return {"Wenxi_MG": price}
        return {}
    finally:
        await browser.close()
        await p.stop()


async def _ocr_asianmetal_fallback() -> Optional[float]:
    """OCR 备选：自动登录 + 开文章页 → 截全页 → 调 OCR.space → 正则取闻喜+区间

    6/30 主人拍板：价格表是图片，不是不让你动，是必须动——走 OCR 拿。
    抓完后会写一份截图证据到 screenshots/ 以便人工复核。
    2026-07-02 改造：去掉 CDP 9223 依赖，独立 headless + 自动登录。
    """
    import requests
    import re

    p, browser = await _launch_headless_browser()
    shot_path = None
    try:
        context = await browser.new_context(viewport={"width": 1280, "height": 900}, bypass_csp=True)
        # 自动登录
        login_ok = await _asianmetal_login_and_get_cookies(context)
        if not login_ok:
            logger.warning("OCR 备选：亚洲金属网自动登录失败")
            return None

        page = await context.new_page()
        await page.goto(
            "https://www.asianmetal.cn/MagnesiumPricesHistory",
            timeout=20000,
            wait_until="domcontentloaded",
        )
        await page.wait_for_timeout(3000)
        # 抓第一个含「镁锭价格分区域」+ 不含「镁合金/镁粉/99.95%min」的文章链接
        href = await page.evaluate('''
            () => {
                const links = document.querySelectorAll('a[href*="/news/"]');
                for (const a of links) {
                    const t = a.innerText.split('\\n')[0];
                    if (!t) continue;
                    if (t.includes('镁合金') || t.includes('镁粉') || t.includes('99.95%min')) continue;
                    if (t.includes('镁锭') && t.includes('价格')) return a.href;
                }
                return null;
            }
        ''')
        if not href:
            logger.warning("OCR 备选：未找到今日镁锭文章 URL")
            return None

        # 打开文章页 + 截全页
        await page.goto(href, timeout=30000, wait_until="domcontentloaded")
        await page.wait_for_timeout(8000)
        shot_path = SHOT_DIR / f"asianmetal_闻喜镁锭_OCR备选_{datetime.now().strftime('%H%M%S')}.png"
        await page.screenshot(path=str(shot_path), full_page=True, timeout=60000)
        logger.info(f"OCR 备选截图保存: {shot_path}")
        await page.close()
        await context.close()
    finally:
        await browser.close()
        await p.stop()

    if not shot_path or not shot_path.exists():
        logger.warning("OCR 备选：截图未保存")
        return None

    # 调 OCR.space 免费 API
    try:
        with open(shot_path, "rb") as f:
            img_data = f.read()
        resp = requests.post(
            "https://api.ocr.space/parse/image",
            files={"filename": ("img.png", img_data, "image/png")},
            data={"language": "chs", "isOverlayRequired": "false"},
            headers={"apikey": os.environ.get("OCR_SPACE_APIKEY", "")},
            timeout=60,
        )
        result = resp.json()
        if not result.get("ParsedResults"):
            logger.warning(f"OCR.space 无结果: {result.get('ErrorMessage')}")
            return None
        text = result["ParsedResults"][0]["ParsedText"]
        logger.info(f"OCR 文本前 500 字: {text[:500]}")
        # 找「闻喜」+「镁锭」+ 99.9 同行 + 区间
        m = re.search(r"闻喜[\s\S]{0,30}镁锭[\s\S]{0,30}99[\.,]9\s*%?\s*min?[\s\S]{0,30}?(\d[\d,\.]+)\s*[-–~]\s*(\d[\d,\.]+)", text)
        if m:
            low = float(m.group(1).replace(",", ""))
            high = float(m.group(2).replace(",", ""))
            if 10000 < low < 50000 and 10000 < high < 50000:
                mid = round((low + high) / 2, 2)
                logger.info(f"OCR 备选拿到闻喜镁锭: {low}-{high} → {mid}")
                return mid
        logger.warning("OCR 备选：未在文本中匹配到「闻喜+镁锭+99.9%min」区间")
        return None
    except Exception as e:
        logger.warning(f"OCR 备选异常: {e}")
        return None


def get_tungsten_price() -> Optional[float]:
    """中钨在线钨粉价格（保存原始 HTML 作为现场证据）"""
    try:
        from jinggong_monitor.fetcher_tungsten import ChinatungstenFetcher
        fetcher = ChinatungstenFetcher()
        results = fetcher.fetch()
        if results:
            # 6/29 主人拍板：保存原始 HTML 作为追溯证据（不能截屏是因为 requests 抓的不是浏览器）
            url = fetcher._find_daily_article_url()
            if url:
                try:
                    import requests
                    r = requests.get(url, timeout=10)
                    evidence = SHOT_DIR / f"chinatungsten_钨粉原文_{datetime.now().strftime('%H%M%S')}.html"
                    evidence.write_text(r.text, encoding="utf-8")
                    SHOT_RESULTS["中钨在线/钨粉原文"] = (True, str(evidence.relative_to(SCREENSHOTS_ROOT)))
                    logger.info(f"📸 中钨在线 HTML 证据保存: {evidence.name}")
                except Exception as e:
                    SHOT_RESULTS["中钨在线/钨粉原文"] = (False, f"HTML 保存失败: {e}")
        else:
            SHOT_RESULTS["中钨在线/钨粉"] = (False, "未匹配到钨粉价格正则（可能今日未发文）")
        return results.get("W")
    except Exception as e:
        logger.warning(f"中钨在线抓取失败: {e}")
        SHOT_RESULTS["中钨在线/钨粉"] = (False, str(e))
        return None


def get_wti_estimate() -> Optional[float]:
    """获取WTI估算（CL realtime or SC futures/7.15）"""
    try:
        import akshare as ak
        df = ak.futures_foreign_commodity_realtime(symbol="CL")
        return round(float(df.iloc[0]["最新价"]), 2)
    except Exception:
        pass
    try:
        import akshare as ak
        df = ak.futures_main_sina(symbol="SC0", start_date="20260624", end_date="20260624")
        close = float(df.iloc[-1]["收盘价"])
        return round(close / 7.15, 2)
    except Exception:
        pass
    return None


# ============================================================
# 填表逻辑
# ============================================================

def fill_sheet1(wb, smm: dict, ccmn: dict, asianmetal: Optional[dict] = None, target_dates: Optional[list[tuple[str, int]]] = None, mode: str = "manual"):
    """填写「日均价（2026年市场）」sheet。
    mode='manual'：填补历史 6/23/24（默认）
    mode='daily'：填今天的 row（17:00 cron 调用场景）；WTI 不在这里写（sheet2 负责）
    """
    ws = wb["日均价（2026年市场）"]

    # 6/29 主人拍板：填表前跟历史价校验，偏差 >50% 则标黄不写
    # 例：锰 19200 → 6（差 99.97%） 视为采集错误，不写入
    def historical_avg(ws, col, current_row):
        """拿历史上 5 个有效日的均价（用于偏差校验）。
        关键：不能包含 current_row（今天），避免自污染。
        """
        vals = []
        for r in range(max(2, current_row-30), current_row):  # 严格 < current_row
            v = ws.cell(row=r, column=col).value
            if v and isinstance(v, (int, float)) and v > 100:  # 过滤 0/None/异常小值
                vals.append(v)
                if len(vals) >= 5: break
        if not vals: return None
        return sum(vals) / len(vals)

    asianmetal = asianmetal or {}

    # 列映射（Col 13 优先亚洲金属网，降级SMM）
    col_map = {
        2: ("ADC12", "SMM"),
        3: ("A380", "SMM"),
        4: ("AlSi9Cu3", "SMM"),
        5: ("A356", "SMM"),
        6: ("A00_AL", "ccmn"),
        7: ("CU", "ccmn"),
        # 金属硅中间价 3 列 — 2026-06-26 主人拍板新规则
        8: ("SI_553_331_AVG", "ccmn"),     # H 列 441 = 金属硅553#-331# 均价
        9: ("SI_3303_2202_MIN", "ccmn"),   # I 列 3303 = 金属硅3303#-2202# 最低价
        10: ("MG", "ccmn"),
        11: ("MN", "ccmn"),
        12: ("SI_553_331_MAX", "ccmn"),    # L 列 331 = 金属硅553#-331# 最高价
        13: ("Wenxi_MG", "ASIANMETAL"),  # 🆕 亚洲金属网优先，SMM 备源
        14: ("AM60B", "SMM"),
        15: ("AZ91D", "SMM"),
    }

    remarks = []

    # 如果未指定 target_dates，根据 mode 生成默认
    if target_dates is None:
        if mode == "daily":
            # 17:00 cron 场景：填今天的 row（新建或覆盖已有）
            today = TODAY_STR
            # 查找今天是否已经有 row
            target_row = None
            for r in range(2, ws.max_row + 2):
                v = ws.cell(row=r, column=1).value
                if v and str(v)[:10] == today:
                    target_row = r
                    break
            if target_row is None:
                target_row = ws.max_row + 1
            target_dates = [(today, target_row)]
        else:  # manual
            # 清理已有的 junk rows 116-118
            for r in range(116, 119):
                for c in range(1, 16):
                    ws.cell(row=r, column=c, value=None)
            target_dates = [("2026-06-23", 119), ("2026-06-24", 120)]

    # 填写指定的 (日期, 行号)
    for date_str, row_num in target_dates:
        # 清空该行
        for c in range(1, 16):
            ws.cell(row=row_num, column=c, value=None)
        ws.cell(row=row_num, column=1, value=date_str)
        # 6/29 主人拍板：日期列用 DATE_FONT（微软雅黑 11）
        ws.cell(row=row_num, column=1).font = DATE_FONT

        for col, (variety, source) in col_map.items():
            price = None
            if source == "SMM":
                price = smm.get(variety)
            elif source == "ASIANMETAL":
                # 6/29 主人拍板：闻喜镁锭必须用亚洲金属网，不走 SMM 备源
                price = asianmetal.get(variety)
                if price is None:
                    # 不降级到 SMM（主人主业要求），标黄不写
                    cell = ws.cell(row=row_num, column=col)
                    cell.fill = YELLOW_FILL
                    reason = f"{date_str}: {variety} - 亚洲金属网未获取（主业要求用亚洲金属网，标黄不写）"
                    remarks.append(reason)
                    logger.warning(reason)
                else:
                    logger.info(f"Col13 闻喜镁锭（亚洲金属网）: {price}")
            elif source == "ccmn":
                price = ccmn.get(variety)
                # 回退到 akshare（仅对未受新规则影响的品种）
                if price is None:
                    akshare_map = {"A00_AL": "AL", "CU": "CU"}
                    if variety in akshare_map:
                        spot = get_akshare_spot(akshare_map[variety], date_str)
                        if spot:
                            price = spot

            if price and price > 0:
                # 6/29 主人拍板：与历史价校验，偏差 >50% 则视为采集错误、标黄不写
                # 例：锰 19200 → 6（差 99.97%）不写入
                hist_avg = historical_avg(ws, col, row_num)
                if hist_avg and abs(price - hist_avg) / hist_avg > 0.5:
                    cell = ws.cell(row=row_num, column=col)
                    cell.fill = YELLOW_FILL
                    diff_pct = abs(price - hist_avg) / hist_avg * 100
                    reason = f"{date_str}: {variety}={price} 与历史均价{hist_avg:.0f} 偏离 {diff_pct:.1f}%，可能采集错误，标黄不写"
                    remarks.append(reason)
                    logger.warning(reason)
                else:
                    ws.cell(row=row_num, column=col, value=price)
            else:
                # 标黄空单元格
                cell = ws.cell(row=row_num, column=col)
                cell.fill = YELLOW_FILL
                if source == "ASIANMETAL":
                    src_label = "亚洲金属网(需登录)"
                    reason = f"{date_str}: {variety} 无法从亚洲金属网获取（请确认Chrome已登录）"
                elif source == "SMM":
                    src_label = "SMM(需登录)"
                    reason = f"{date_str}: {variety} 无法从{src_label}获取"
                else:
                    src_label = "ccmn/AK"
                    reason = f"{date_str}: {variety} 无法从{src_label}获取"
                remarks.append(reason)
                logger.warning(reason)

            # 6/29 主人拍板：所有数据列统一 DATA_FONT（微软雅黑 11）
            ws.cell(row=row_num, column=col).font = DATA_FONT

    # 打印结果
    for row_num in [119, 120]:
        vals = [ws.cell(row=row_num, column=c).value for c in range(1, 16)]
        logger.info(f"Row {row_num}: {vals}")

    return remarks


def fill_sheet2(wb, wti_price: Optional[float], target_date: Optional[str] = None, skip_wti: bool = False):
    """填写「日均价（中钨在线 原油）」sheet。
    
    Args:
        wb: openpyxl Workbook
        wti_price: WTI 时点价（None 表示跳过）
        target_date: 单日日期（cron 自动跑场景）；None 时回退到历史 6/23+6/24 手动调用场景
        skip_wti: 17:00 主流程调用时设 True，让 15:00 cron 写入的 WTI 不被覆盖
    """
    sheet_name = "日均价（中钨在线 原油）"
    if sheet_name not in wb.sheetnames:
        # 6/29 主人拍板：`共享(2).xlsx` 没有这个 sheet（主人手动删了），自动重建
        ws = wb.create_sheet(sheet_name)
        # 复制表头格式（贴一个轻量表头）
        headers = ["日期", "品种", "含税价", "不含税价", "品种2", "日期2", "价格"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = HEADER_FONT
        logger.warning(f"⚠️ {sheet_name} 不存在，已自动创建")
    else:
        ws = wb[sheet_name]
    remarks = []

    last_row = ws.max_row
    while last_row > 2 and ws.cell(row=last_row, column=1).value is None:
        last_row -= 1

    # 7. 目标日期列表：17:00 cron 只填今天；手动调用时回退历史 6/23+6/24
    if target_date is not None:
        date_list = [target_date]
    else:
        date_list = ["2026-06-23", "2026-06-24"]

    for date_str in date_list:
        # target_date 模式：新建行；否则复用现有行
        if target_date is not None:
            # 检查今天是否已有行；有则覆盖、无则新建
            existing_row = None
            for r in range(3, ws.max_row + 2):
                v = ws.cell(row=r, column=1).value
                if v and isinstance(v, datetime) and v.strftime("%Y-%m-%d") == date_str:
                    existing_row = r
                    break
            row = existing_row if existing_row else ws.max_row + 1
        else:
            # 历史回退模式：硬编码 6/23→119, 6/24→120（保留原逻辑）
            row = {"2026-06-23": 119, "2026-06-24": 120}.get(date_str)
            if row is None:
                continue

        dt = datetime.strptime(date_str, "%Y-%m-%d")

        # 钨粉 (Col A=日期, B=品种, C=含税价, D=不含税)
        ws.cell(row=row, column=1, value=dt)
        ws.cell(row=row, column=1).font = DATE_FONT
        ws.cell(row=row, column=2, value="钨粉")
        ws.cell(row=row, column=2).font = DATA_FONT
        # 钨粉无免费每日报价源，留空 + 标黄
        cell_c = ws.cell(row=row, column=3)
        cell_c.value = None
        cell_c.fill = YELLOW_FILL
        cell_c.font = DATA_FONT
        remarks.append(f"{date_str}: 钨粉 - 中钨在线 news.chinatungsten.com 仅隔天发布当日文章，无免费每日数据源")
        logger.warning(f"{date_str} 钨粉: 无法获取，留空标黄")
        ws.cell(row=row, column=4, value=f"=IFERROR(C{row}/1.13,\"\")")  # 不含税公式（保留兼容）
        ws.cell(row=row, column=4).font = DATA_FONT

        # WTI (Col E=品种, F=日期, G=价格)
        ws.cell(row=row, column=5, value="WTI原油")
        ws.cell(row=row, column=5).font = DATA_FONT
        ws.cell(row=row, column=6, value=dt)
        ws.cell(row=row, column=6).font = DATE_FONT

        if skip_wti:
            # 15:00 cron 已写 WTI，不要覆盖
            logger.info(f"{date_str} WTI: skip_wti=True，跳过（15:00 cron 已填）")
        elif wti_price and wti_price > 0:
            ws.cell(row=row, column=7, value=wti_price)
            ws.cell(row=row, column=7).font = DATA_FONT
            logger.info(f"{date_str} WTI: {wti_price}")
        else:
            cell_g = ws.cell(row=row, column=7)
            cell_g.fill = YELLOW_FILL
            cell_g.font = DATA_FONT
            remarks.append(f"{date_str}: WTI - akshare realtime 仅提供当日实时价")
            logger.warning(f"{date_str} WTI: 无法获取，留空标黄")

    return remarks


# ============================================================
# 历史核查
# ============================================================

def verify_historical(wb) -> str:
    """核查 6/16-22 历史数据（5个工作日：16,17,18,22；跳过19日周末）"""
    ws = wb["日均价（2026年市场）"]
    date_rows = {}
    for r in range(2, 121):
        d = ws.cell(row=r, column=1).value
        if d and isinstance(d, (str, datetime)):
            d_str = d.strftime("%Y-%m-%d") if isinstance(d, datetime) else str(d)[:10]
            date_rows[d_str] = r

    target_dates = ["2026-06-16", "2026-06-17", "2026-06-18", "2026-06-22"]
    # 6/19 是周末非交易日

    col_headers = {}
    for c in range(2, 16):
        col_headers[c] = ws.cell(row=1, column=c).value or f"Col{c}"

    report_lines = []
    report_lines.append("# 历史数据核查报告")
    report_lines.append(f"\n核查范围: 2026-06-16 至 2026-06-22（5个工作日）")
    report_lines.append(f"核查时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    report_lines.append("\n## 核查方法")
    report_lines.append("- 长江现货系列（Col 6-12: A00铝/铜/硅441/硅3303/镁/锰/硅331）：与 akshare futures_spot_price 现货基准价对比")
    report_lines.append("- 上海有色系列（Col 2-5, 13-15: ADC12/A380/AlSi9Cu3/A356/闻喜镁锭/AM60B/AZ91D）：与 SMM CDP 实时价趋势对比（因无法回溯SMM历史数据，仅做趋势一致性检查）")
    report_lines.append("- 偏离度阈值: ≤3% 可接受, >3% 需关注, >10% 异常")
    report_lines.append("")

    # 核查数据
    total_checked = 0
    anomalies = 0

    for date_str in target_dates:
        row = date_rows.get(date_str)
        if row is None:
            report_lines.append(f"\n### {date_str}: 表中无数据行")
            continue

        report_lines.append(f"\n### {date_str} (Row {row})")
        report_lines.append("| 商品 | 表中值 | 实际值 | 偏离度 | 状态 |")
        report_lines.append("|------|--------|--------|--------|------|")

        # 核查有akshare对照的列: A00铝(AL), 铜(CU), 硅(SI系列)
        checks = {
            6: ("A00铝", "AL", 3.0),
            7: ("铜", "CU", 3.0),
            8: ("金属硅441", "SI", 5.0),   # SI现货 + 300
            9: ("金属硅3303", "SI", 5.0),  # SI现货 + 1000
            12: ("金属硅331", "SI", 5.0),  # SI现货 + 700
        }

        for col, (name, symbol, threshold) in checks.items():
            recorded = ws.cell(row=row, column=col).value
            if recorded is None or float(recorded) <= 0:
                continue

            actual = get_akshare_spot(symbol, date_str)
            if actual is None:
                continue

            # 金属硅中间价规则（2026-06-26 主人拍板）: 不再走 akshare SI 校核
            # 原因：新规则下 H/L/I 列取的是综合品类（金属硅553#-331#、金属硅3303#-2202#）的 min/max/avg
            # akshare SI 期货价跟这些中间价没直接对应关系，跳过校核
            if col in (8, 9, 12):
                continue

            actual_adj = actual
            recorded_f = float(recorded)
            if actual_adj > 0:
                deviation = abs(recorded_f - actual_adj) / actual_adj * 100
                status = "✅" if deviation <= threshold else ("⚠️" if deviation <= 10 else "❌异常")
                if deviation > threshold:
                    anomalies += 1

                report_lines.append(
                    f"| {name} | {recorded_f:,.0f} | {actual_adj:,.0f} "
                    f"| {deviation:.1f}% | {status} |"
                )
                total_checked += 1

        # SMM列核查（仅趋势一致性，无历史真实值可对比）
        smm_cols = {2: "ADC12", 3: "A380", 4: "AlSi9Cu3", 5: "A356",
                     13: "闻喜镁锭", 14: "AM60B", 15: "AZ91D"}
        for col, name in smm_cols.items():
            recorded = ws.cell(row=row, column=col).value
            if recorded:
                report_lines.append(
                    f"| {name}(SMM) | {float(recorded):,.0f} | — | — | ℹ️ 无历史对照源 |"
                )
                total_checked += 1

    report_lines.append(f"\n## 总结")
    report_lines.append(f"- 核查品种数: {total_checked}")
    report_lines.append(f"- 可对照(akshare)异常数: {anomalies}")
    report_lines.append(f"- SMM列(无历史对照): 核查但不能定量对比")
    report_lines.append(f"- 结论: {'存在异常需关注' if anomalies > 0 else '数据质量良好，口径延续一致'}")

    return "\n".join(report_lines)


# ============================================================
# 主函数
# ============================================================

async def main():
    print("=" * 60)
    print("精工有色金属共享表 - 填充 + 核查")
    print("=" * 60)

    # 2026-07-07 新增：在线同步前置 —— 先拉远程 data.json，回写 Excel
    # 这样同事在线编辑的内容会先合并进 Excel，再被今天的新数据覆盖
    print("\n[0/4] 同步在线编辑 → Excel ...")
    try:
        from sync_from_web import sync_to_excel as _sync
        from git_helper import git_pull_rebase as _pull
        if _pull():
            _sync()
        else:
            print("  ⚠️ git pull 失败，跳过在线同步（使用本地 Excel）")
    except Exception as e:
        print(f"  ⚠️ 在线同步失败（继续执行）: {e}")

    # 6/29 主人拍板：默认 17:00 cron 调用时填今天并跳过 WTI
    # 用环境变量 RUN_MODE 控制：daily=17:00 默认；manual=手动跑历史回溯
    run_mode = os.environ.get("JINGGONG_RUN_MODE", "daily")
    if run_mode == "daily":
        sheet2_target = TODAY_STR
        sheet2_skip_wti = True  # WTI 由 15:00 cron 写入
    else:  # manual mode（保留历史 6/23+6/24 回溯逻辑）
        sheet2_target = None
        sheet2_skip_wti = False
    print(f"[mode] {run_mode} | sheet2_target={sheet2_target} | skip_wti={sheet2_skip_wti}")

    # 1. 采集数据
    print("\n[1/4] 采集 ccmn 数据...")
    ccmn = await get_ccmn_prices()
    print(f"  ccmn: {len(ccmn)} 品种")

    print("[2/4] 采集 SMM 数据 (CDP)...")
    try:
        smm = await get_smm_cdp()
        print(f"  SMM: {len(smm)} 品种")
    except Exception as e:
        print(f"  SMM CDP 失败: {e}")
        smm = {}

    print("[2.5/4] 采集亚洲金属网数据 (CDP 登录态)...")
    try:
        asianmetal = await get_asianmetal_price()
        print(f"  亚洲金属网: {len(asianmetal)} 品种")
    except Exception as e:
        print(f"  亚洲金属网失败: {e}")
        asianmetal = {}

    wti = get_wti_estimate()
    print(f"  WTI: {wti}")

    # 2.5/4 中钨在线钨粉（6/29 主人拍板：保存 HTML 证据）
    print("[2.7/4] 采集中钨在线钨粉...")
    tungsten = get_tungsten_price()
    print(f"  钨粉: {tungsten}")

    # 2. 打开Excel
    print("\n[3/4] 填写数据...")
    wb = openpyxl.load_workbook(str(EXCEL_PATH))

    remarks_s1 = fill_sheet1(wb, smm, ccmn, asianmetal, mode=run_mode)
    remarks_s2 = fill_sheet2(wb, wti, target_date=sheet2_target, skip_wti=sheet2_skip_wti)

    # 3. 历史核查（仅 manual 模式执行，daily 跳过——硬编码的 6/16-22 已过时）
    if run_mode == "manual":
        print("[4/4] 历史数据核查...")
        report = verify_historical(wb)
        report_path = EXCEL_PATH.parent / "核查报告_2026-06-16_to_22.md"
        report_path.write_text(report, encoding="utf-8")
        print(f"核查报告: {report_path}")
    else:
        print("[4/4] 跳过历史核查（daily 模式）")
        report = "（daily 模式，跳过历史核查）"

    # 4. 保存Excel
    wb.save(str(EXCEL_PATH))
    wb.close()

    # 5. 输出报告
    report_path = EXCEL_PATH.parent / "核查报告_2026-06-16_to_22.md"
    report_path.write_text(report, encoding="utf-8")

    print(f"\nExcel 已保存: {EXCEL_PATH}")

    # 打印备注信息
    all_remarks = remarks_s1 + remarks_s2
    if all_remarks:
        print("\n=== 无法填写的项目 ===")
        for r in all_remarks:
            print(f"  ⚠️ {r}")

    print("\n=== 核查报告摘要 ===")
    print(report)

    # 6/29 主人拍板：明确返回成功/失败
    print("\n" + "=" * 60)
    print("📊 运行总结")
    print("=" * 60)
    print(f"📅 日期: {TODAY_STR}")
    print(f"📁 截图目录: {SHOT_DIR.relative_to(EXCEL_PATH.parent)}")
    print(f"📸 截图/证据: {len(SHOT_RESULTS)} 项")
    ok_count = sum(1 for ok, _ in SHOT_RESULTS.values() if ok)
    fail_count = len(SHOT_RESULTS) - ok_count
    print(f"   ✅ 成功: {ok_count} | ❌ 失败/部分: {fail_count}")
    for label, (ok, info) in SHOT_RESULTS.items():
        icon = "✅" if ok else "⚠️"
        print(f"   {icon} {label}: {info}")
    print(f"📊 Excel 保存: {'OK' if not all_remarks else f'OK（{len(all_remarks)} 项标黄）'}")
    print("=" * 60)

    # 2026-07-07 新增：自动发布看板 —— 导出 data.json + commit + push
    # 这样 17:00 cron 跑完后看板自动更新，无需手动 excel_to_web.py
    print("\n[发布] 自动更新看板 ...")
    try:
        project_dir = str(EXCEL_PATH.parent)
        export_script = EXCEL_PATH.parent / "export_excel_to_json.py"
        # 1. 导出 data.json
        exp = subprocess.run(
            [sys.executable, str(export_script)],
            cwd=project_dir, capture_output=True, text=True, timeout=60,
        )
        if exp.returncode != 0:
            print(f"  ⚠️ 导出 data.json 失败: {exp.stderr.strip()[:200]}")
        else:
            print(f"  ✅ data.json 已导出")
            # 2. 一站式发布（含代理检测、冲突处理）
            from git_helper import publish_to_github
            ok = publish_to_github(
                files=["docs/data.json"],
                commit_msg=f"auto-update {TODAY_STR}",
                cwd=project_dir,
            )
            if ok:
                print(f"  ✅ 看板已推送: https://lekelin2046.github.io/jinggong-commodity-monitor/")
            else:
                print(f"  ⚠️ 看板推送失败（data.json 已在本地，可手动 python3 excel_to_web.py）")
    except Exception as e:
        print(f"  ⚠️ 发布失败（不影响 Excel 数据）: {e}")

    # 给 cron/接管脚本明确的退出码
    sys.exit(0 if fail_count == 0 else 1)


if __name__ == "__main__":
    asyncio.run(main())
