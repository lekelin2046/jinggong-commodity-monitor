#!/usr/bin/env python3
"""从 中钨在线 文章 HTML 手动回填 钨粉(W) 价格。

使用场景：中钨在线网站宕机/抓取失败时，主人从微信专辑页打开当日文章，
把页面保存为 HTML 文件，运行本脚本一键回填到 Excel 并发布。

用法（预览）：
    python3 backfill_tungsten_from_article.py article.html --date 2026-09-04
用法（确认写入并推送）：
    python3 backfill_tungsten_from_article.py article.html --date 2026-09-04 --apply

提取规则：
    优先解析文章内嵌 HTML 报价表：产品名称=钨粉、单位=元/千克 那一行的
    「中钨在线报价」列。解析不到则返回失败，不会用散文正则兜底——避免再犯
    2026-09-03 把正文里的昨日价当成当日价的错误。
"""

import argparse
import subprocess
import sys
from datetime import datetime
from pathlib import Path

from jinggong_monitor.fetcher_tungsten import _extract_w_from_table

# 项目文件路径
EXCEL = Path("2026年有色金属市场价格.xlsx")
SHEET_NAME = "日均价（2026年市场）"
W_COL = 16  # P 列


def find_date_row(ws, date_str: str):
    """在 Excel 第一列找到目标日期所在行。"""
    target = datetime.strptime(date_str, "%Y-%m-%d")
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v == target:
            return r
    return None


def _norm_old(old):
    """把 Excel 旧值规范化为 float 或 None。"""
    if old is None:
        return None
    if isinstance(old, bool):
        return None
    if isinstance(old, (int, float)):
        return float(old)
    if isinstance(old, str):
        s = old.strip()
        if s in ("", "—", "-", "——"):
            return None
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None
    return None


def main():
    parser = argparse.ArgumentParser(description="从 中钨在线 文章 HTML 回填钨粉价格")
    parser.add_argument("html", help="文章 HTML 文件路径")
    parser.add_argument("--date", required=True, help="回填日期，格式 YYYY-MM-DD")
    parser.add_argument("--apply", action="store_true", help="确认写入 Excel、重导 data.json 并推送 GitHub Pages")
    args = parser.parse_args()

    html_path = Path(args.html)
    if not html_path.exists():
        print(f"文件不存在: {html_path}", file=sys.stderr)
        sys.exit(1)

    html = html_path.read_text(encoding="utf-8")
    price = _extract_w_from_table(html)
    if price is None:
        print(
            "未从 HTML 报价表中解析到 钨粉 价格。\n"
            "请确认 HTML 里包含「产品名称=钨粉、单位=元/千克」的报价表。",
            file=sys.stderr,
        )
        sys.exit(1)

    print(f"[{args.date}] 钨粉（来自报价表）= {price:.2f} 元/千克")

    if not args.apply:
        print("预览完成。确认无误后加 --apply 写入 Excel 并推送。")
        return

    import openpyxl

    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET_NAME]
    row = find_date_row(ws, args.date)
    if row is None:
        print(f"Excel 中未找到日期 {args.date}", file=sys.stderr)
        sys.exit(1)

    old_value = ws.cell(row, W_COL).value
    ws.cell(row, W_COL).value = price
    wb.save(EXCEL)
    print(f"Excel row {row}: 钨粉 {old_value} -> {price:.2f}")

    # 重导 data.json
    subprocess.run([sys.executable, "export_excel_to_json.py"], check=True)
    print("docs/data.json 已重导")

    # 变更留痕
    import changelog as cl
    diffs = [
        {
            "date_row": args.date,
            "code": "W",
            "old": _norm_old(old_value),
            "new": float(price),
        }
    ]
    n = cl.record_changes(
        diffs,
        source=cl.SOURCE_MANUAL_XLSX,
        editor="—",
        commit=cl.current_commit_sha(),
    )
    print(f"docs/changelog.json 写入 {n} 条")

    # 推送（只推 data.json + changelog.json；Excel 被 .gitignore 忽略）
    import git_helper as gh

    ok = gh.publish_to_github(
        ["docs/data.json", "docs/changelog.json"],
        f"手动回填 {args.date} 钨粉={price:.2f} 元/千克（来自中钨在线文章 HTML 报价表）",
    )
    if not ok:
        print("GitHub 推送失败，本地修改和 commit 已保留", file=sys.stderr)
        sys.exit(1)
    print("GitHub Pages 已更新")


if __name__ == "__main__":
    main()
