#!/usr/bin/env python3
"""
同步线上 data.json → 本地 Excel

同事在 editor.html 在线编辑后数据已提交到 GitHub，
运行此脚本：git pull → 读 data.json → 写 Excel → 本地表同步。

2026-07-07 修复：
1. Excel 路径统一为「共享(2).xlsx」（与 cron 一致）
2. null 值清空 Excel 单元格（之前跳过 null，导致线上清空无法回写）
3. 自动新建 Excel 没有的日期行（之前跳过，导致线上补录历史日期失败）
4. 保留 datetime 格式（避免 Excel 日期格式混乱）

用法: python3 sync_from_web.py
"""

import sys, os, json, datetime, subprocess
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

SCRIPT_DIR = Path(__file__).parent
# 2026-07-07：与 cron 路径统一（fill_and_verify.py 写共享(2).xlsx）
EXCEL_PATH = SCRIPT_DIR / "2026年有色金属市场价格共享(2).xlsx"
JSON_PATH = SCRIPT_DIR / "docs" / "data.json"
SHEET_NAME = "日均价（2026年市场）"

COLUMN_MAP = {
    2: "ADC12", 3: "A380", 4: "AlSi9Cu3", 5: "A356",
    6: "A00_AL", 7: "CU", 8: "SI_441", 9: "SI_3303",
    10: "MG", 11: "MN", 12: "SI_331",
    13: "Wenxi_MG", 14: "AM60B", 15: "AZ91D",
    16: "W", 17: "WTI",
}
CODE_TO_COL = {v: k for k, v in COLUMN_MAP.items()}

DATE_FONT = Font(name="微软雅黑", size=11, bold=False)


def git_pull():
    """拉取最新代码（含同事在线编辑的 data.json）"""
    print("→ git pull --rebase ...")
    from git_helper import git_pull_rebase
    ok = git_pull_rebase(str(SCRIPT_DIR))
    if ok:
        print("  ✅ 拉取成功")
    else:
        print("  ⚠️ 拉取失败（可能需要代理或网络不可用）")
    return ok


def _parse_excel_date(val):
    """将 Excel A 列的值解析为 'YYYY-MM-DD' 字符串，无法解析返回 None"""
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        s = val.strip()
        if not s or s.startswith("备注"):
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None


def sync_to_excel():
    """读 data.json → 写入 Excel
    
    - null 值清空对应单元格
    - Excel 没有的日期行自动新建
    """
    if not JSON_PATH.exists():
        print(f"ERROR: {JSON_PATH} 不存在", file=sys.stderr)
        return False

    with open(JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    # 构建 date_str -> row_number 映射
    date_rows = {}
    for r in range(2, ws.max_row + 1):
        ds = _parse_excel_date(ws.cell(r, 1).value)
        if ds:
            date_rows[ds] = r

    updated = 0
    cleared = 0
    new_rows = 0

    for ds, row_data in sorted(data.get("data", {}).items()):
        # 新建 Excel 没有的日期行
        if ds not in date_rows:
            new_row = ws.max_row + 1
            # 写日期（保持字符串格式，与 fill_and_verify.py 一致）
            ws.cell(row=new_row, column=1, value=ds).font = DATE_FONT
            date_rows[ds] = new_row
            new_rows += 1
            print(f"  ➕ 新建日期行 {ds} (row {new_row})")

        r = date_rows[ds]
        for code, val in row_data.items():
            col = CODE_TO_COL.get(code)
            if col is None:
                continue
            existing = ws.cell(r, col).value
            if val is None:
                # null → 清空单元格（之前会跳过，导致线上清空无法回写）
                if existing not in (None, "", "——", "—", "-"):
                    ws.cell(r, col).value = None
                    cleared += 1
            else:
                if existing != val:
                    ws.cell(r, col).value = val
                    updated += 1

    wb.save(EXCEL_PATH)
    print(f"✓ 写入完成: 更新 {updated} 个单元格 | 清空 {cleared} 个 | 新建 {new_rows} 行")
    return True


if __name__ == "__main__":
    print(f"=== 同步线上数据 → 本地 Excel ===\n")
    if not git_pull():
        sys.exit(1)
    if not sync_to_excel():
        sys.exit(1)
    print(f"\n✓ 完成！Excel 已更新：{EXCEL_PATH.name}")
    print(f"  打开 Excel 确认后，运行 python3 excel_to_web.py 推回线上")
