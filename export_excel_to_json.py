#!/usr/bin/env python3
"""
Excel -> JSON 导出脚本
读取 2026年有色金属市场价格.xlsx 的"日均价（2026年市场）"sheet，
输出 docs/data.json 供看板 HTML 动态加载。

用法: python3 export_excel_to_json.py [excel_path]
默认: 2026年有色金属市场价格.xlsx
"""

import sys
import os
import json
import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ===== 配置 =====
SCRIPT_DIR = Path(__file__).parent
# 2026-07-07：统一用「共享(2).xlsx」（cron 实际写入的文件），避免双 Excel 分叉
DEFAULT_EXCEL = SCRIPT_DIR / "2026年有色金属市场价格共享(2).xlsx"
OUTPUT_JSON = SCRIPT_DIR / "docs" / "data.json"
SHEET_NAME = "日均价（2026年市场）"

# Excel 列 -> 品种代码映射 (1-indexed)
COLUMN_MAP = {
    2:  "ADC12",
    3:  "A380",
    4:  "AlSi9Cu3",
    5:  "A356",
    6:  "A00_AL",
    7:  "CU",
    8:  "SI_441",
    9:  "SI_3303",
    10: "MG",
    11: "MN",
    12: "SI_331",
    13: "Wenxi_MG",
    14: "AM60B",
    15: "AZ91D",
    16: "W",
    17: "WTI",
}


def parse_value(val):
    """将 Excel 单元格值转为数字或 None"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s in ("", "——", "-", "—"):
        return None
    # 尝试去掉逗号
    s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def export(excel_path: str, output_path: str):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    ws = wb[SHEET_NAME]
    data = {}
    row_count = 0
    latest_date = None

    for r in range(2, ws.max_row + 1):
        date_val = ws.cell(r, 1).value
        if date_val is None:
            continue
        # 接受 datetime 和字符串两种日期格式（fill_and_verify.py 写字符串，历史行是 datetime）
        if isinstance(date_val, datetime.datetime):
            if date_val.year != 2026:
                continue
            date_str = date_val.strftime("%Y-%m-%d")
        elif isinstance(date_val, str):
            s = date_val.strip()
            if not s or s.startswith("备注"):
                continue
            # 尝试多种格式
            parsed = None
            for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
                try:
                    parsed = datetime.datetime.strptime(s, fmt)
                    break
                except ValueError:
                    continue
            if parsed is None or parsed.year != 2026:
                continue
            date_str = parsed.strftime("%Y-%m-%d")
        else:
            continue

        day_data = {}
        has_any = False

        for col, code in COLUMN_MAP.items():
            val = parse_value(ws.cell(r, col).value)
            if val is not None:
                day_data[code] = val
                has_any = True

        if has_any:
            data[date_str] = day_data
            row_count += 1
            if latest_date is None or date_str > latest_date:
                latest_date = date_str

    # 构建输出对象
    output = {
        "last_updated": latest_date or datetime.date.today().isoformat(),
        "total_days": row_count,
        "varieties": list(COLUMN_MAP.values()),
        "data": data,
    }

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"OK: {row_count} days exported")
    print(f"  Source: {excel_path}")
    print(f"  Output: {output_path}")
    print(f"  Latest: {latest_date}")
    print(f"  Size: {os.path.getsize(output_path)} bytes")


if __name__ == "__main__":
    excel_path = sys.argv[1] if len(sys.argv) > 1 else str(DEFAULT_EXCEL)
    if not os.path.exists(excel_path):
        print(f"ERROR: File not found: {excel_path}", file=sys.stderr)
        sys.exit(1)
    export(excel_path, str(OUTPUT_JSON))
